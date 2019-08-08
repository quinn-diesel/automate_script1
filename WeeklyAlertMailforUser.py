


__author__ = 'SrikanthReddy Rokkam, Shirsendu Sarkar'
__copyright__ = 'NavTech India Pvt.Ltd.,'
import smtplib
import os
from openpyxl import load_workbook
import datetime
from smtplib import SMTPRecipientsRefused, SMTPServerDisconnected
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage

from tkMessageBox import *
from collections import OrderedDict
import ConfigParser
from bokeh.models import ColumnDataSource, ranges, LabelSet,LegendItem,Legend
from bokeh.plotting import figure, show, output_file,save

from bokeh.io import export_png

config = ConfigParser.ConfigParser()
confPath = os.path.abspath(os.path.dirname(__file__))+'\config.ini'
config.readfp(open(confPath))

from selenium import webdriver
PhantomJsPath = config.get('PhantomJsPath', 'path')

driver =webdriver.PhantomJS(executable_path = PhantomJsPath)



def getData(infodictnaty):
    table = []
    tabledates=OrderedDict()
    for key, value in infodictnaty.iteritems():
        table.append(r'<tr><th>{}</th><th>{}</th></tr>'.format(key, value))
        tabledates[key]=str(value)
    return ''.join(table),tabledates

def getEmail(person,team):
    email = None
    if team == 'Engg':
        userdeatailPath =os.path.abspath(os.path.dirname(__file__))+'\\UserDeatils.xlsx'
    if team == 'Staff':
        userdeatailPath = os.path.abspath(os.path.dirname(__file__)) + '\\UserDeatils_Staff_Services.xlsx'
    if os.path.exists(userdeatailPath):
        wbuser = load_workbook(userdeatailPath)
        wsuser = wbuser.active
        for rowuser in wsuser.iter_rows(min_row=2, min_col=1, max_col=1, max_row=wsuser.max_row):
            for celluser in rowuser:
                if celluser.value == person:
                    email = wsuser['B'+str(celluser.row)].value
        return email
    else:
        return None
def plotweekgraph(tabledates,person,totalduration):
    print tabledates
    alldates=[]
    allhours=[]
    for i in tabledates:
        alldates.append(str(i))
        if tabledates[i] == '-' or tabledates[i] == 'Missing Out Punch' or tabledates[i] == 'Missing In Punch' or tabledates[i] == 'Absence':
            allhours.append(0)
        else:
            allhours.append(float(str(tabledates[i].split(':')[0])+'.'+str(tabledates[i].split(':')[1])))
    print allhours


    WeeklyGraphpath=os.path.abspath(os.path.dirname(__file__))+'\WeeklyGraphs'
    if not os.path.exists(WeeklyGraphpath):
        os.makedirs(WeeklyGraphpath)
    weeklydatespath= WeeklyGraphpath+'\\'+'WeeklyGraphs_'+alldates[0]+'_To_'+alldates[-1]
    if not os.path.exists(weeklydatespath):
        os.makedirs(weeklydatespath)

    source = ColumnDataSource(dict(x=alldates, y=allhours))

    x_label = "Dates"
    y_label = "Hours"
    title = "week"
    w=0.5
    plot = figure(plot_width=800, plot_height=600, tools="save",
                  x_axis_label=x_label,
                  y_axis_label=y_label,
                  title=title,
                  x_minor_ticks=2,
                  x_range=source.data["x"],
                  y_range=ranges.Range1d(start=0, end=13))

    labels = LabelSet(x='x', y='y', text='y', level='glyph',
                      x_offset=-20, y_offset=0, source=source, render_mode='canvas')

    plot.vbar(source=source, x='x', top='y', bottom=0, width=w)

    li1 = LegendItem(label='Duration :'+totalduration)  # , renderers=[p1.renderers[0]]
    legend1 = Legend(items=[li1], location='top_right', border_line_alpha=0.0)
    plot.add_layout(labels)
    plot.add_layout(legend1)

    image = export_png(plot,filename=weeklydatespath+'\\'+person+'.png',webdriver=driver )
    print image,type(image)
    return image,alldates

def sendmail(userFile, person,infodictnaty,team,mondate,sundate,smtp,totalduration):
    email = getEmail(person,team)
    table,tabledates = getData(infodictnaty)
    pngpath,alldates=plotweekgraph(tabledates,person,totalduration)
    wbuserDeatails = load_workbook(userFile)
    wsuserDeatails = wbuserDeatails.active
    if email != None:
        for rowuserDeatails in wsuserDeatails.iter_rows(min_row=2, min_col=1, max_col=1, max_row=wsuserDeatails.max_row):
            for celluserDeatails in rowuserDeatails:
                if person == celluserDeatails.value:
                    if wsuserDeatails['D' + str(celluserDeatails.row)].value == 'Active':
                        fromaddr = config.get('FromMail', 'from')
                        msg = MIMEMultipart()
                        msg['From'] = fromaddr
                        email='sarkar.santu48@gmail.com'
                        msg['To'] = email
                        msg['Subject'] = 'Weekly Timings Information of %s To %s' % (
                        mondate.strftime('%d-%m-%Y'), sundate.strftime('%d-%m-%Y'))
                        body1 = (r'<p>Dear {},</p><p>           Check the below details</p>'.format(person))
                        body2 = '<p>NOTE: If you need any approval for data modifications mail to respective reporting manger before next 2 days.<p>Regards,<br>HR-Team<br><p style=\"color:red\">This is a system generated email.Please do not reply.'
                        html = """\
                                        <html>
                                        <head>
                                        <style>
                                        table, th, td {
                                            border: 1px solid black;
                                            border-collapse: collapse;
                                            padding: 5px
                                        }
                                        </style>
                                        </head>
                                        <body>
                                        """ + body1 + """
                                        <table style="width:40%">
                                          <tr>
                                            <th  bgcolor="#D5DBDB">Date</th>
                                            <th  bgcolor="#D5DBDB">Time Information</th>
                                          </tr>
                                          """ + table + """</table>""" + body2 + """</body></html>"""
                        pngdata=open(pngpath,"rb").read()
                        image = MIMEImage(pngdata, name='Weekly_Graph_'+alldates[0]+'_To_'+alldates[-1])
                        msg.attach(MIMEText(html, 'html'))
                        msg.attach(image)

                        try:
                            s = smtp  # 207.58.142.98/25
                            s.starttls()
                        except:
                            from FaceRecognitionGui_upt import getSMTP
                            reSmtp = getSMTP()
                            s = reSmtp
                        password = 'Hrams@321$'
                        s.login(fromaddr, password)
                        text = msg.as_string()
                        try:
                            s.sendmail(fromaddr, email, text)
                        except SMTPRecipientsRefused:
                            showerror('', 'incorrect email')
                        except SMTPServerDisconnected:
                            from FaceRecognitionGui_upt import getSMTP
                            reSmtp = getSMTP()
                            reSmtp.sendmail(fromaddr, email, text)
                        print 'Mail Send to ', email,person

def check(userFile, weeklyfile ,team, today,smtp):
    today = datetime.datetime.strptime(today,'%d-%m-%Y')
    mondate = today - datetime.timedelta(days=today.weekday())
    sundate = mondate + datetime.timedelta(days=6)
    wb = load_workbook(weeklyfile)
    ws = wb.active
    totalduration=''
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column-1, max_row=ws.max_row):
        infodictnaty = OrderedDict()
        for cell in row:
            totalduration=ws['I'+str(cell.row)].value
            if cell.column == 'A':
                person =  cell.value
            else :
                date =  ws[cell.column+str(1)].value
                value =  str(cell.value)
                if value == 'None':
                    value = '-'
                infodictnaty[date] = value
        if infodictnaty:
            sendmail(userFile, person,infodictnaty,team,mondate,sundate,smtp,totalduration)
# check('C:\Users\shirsendu sarkar\Desktop\FRSystem_original\UserDeatils.xlsx', 'C:\Users\shirsendu sarkar\Desktop\FRSystem_original\WeeklyTimings\Engineering Services\WeeklyTimingSheet_21-01-2019_To_27-01-2019.xlsx' ,'Engg', '28-01-2019','207.58.142.98/25')