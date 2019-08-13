from openpyxl import load_workbook
import json
import os
import service

EXCEL_PATH = os.getenv("EXCEL_PATH")
SIZE = os.getenv("SIZE")
identity = {
    "username": "Google_106600641818143500761",
    "claims": {
        "email": "ajeya@ljx.com.au",
        "custom:office": "corp",
        "custom:manager": "",
        "custom:team": "engineering",
        "cognito:groups": ["agents-group"],
    },
}


def prepare_payload(address, contacts, primary_agent):
    payload = {
        "arguments": {
            "input": {
                "address": address,
                "contacts": [contacts],
                "primary_agent": primary_agent,
            }
        },
        "identity": identity,
    }
    return payload["arguments"]["input"], payload["identity"]


def main(event, context):

    wb = load_workbook(filename=EXCEL_PATH)
    ws = wb["Sheet3"]

    # want to get every row use 'ws.max_row' instead of 'Limit'
    min_row = 2
    max_row = int(SIZE) + int(min_row) - 1
    for row in ws.iter_rows(
        min_row=min_row, min_col=1, max_row=max_row, max_col=8
    ):  # want to get every row use 'ws.max_row'
        lst = []
        for cell in row:

            lst.append(cell.value)

        # lst[0]=address value
        # lst[1]=contact value
        # lst[2]=agent value
        # lst[3]=bed value
        # lst[4]=bath value
        # lst[5]=car value
        # lst[6]=headline value
        # lst[7]=body value

        if lst[0] != None and lst[1] != None and lst[2] != None:
            # listing=service.get_listing('a10a7f1b-d26f-409e-aaa2-511b392350db',identity)
            # print(listing)
            # service.update_stage(
            #     listing["id"], listing["stage_code"].upper(),identity
            # )
            # print(1/0)
            args, identity = prepare_payload(
                json.loads(lst[0]),
                json.loads(lst[1]),
                json.loads(lst[2])["primaryAgent"],
            )
            # creating a listing in agent_portal
            listing = service.create_listing(args, identity)
            print('listing',listing)
            print('id',listing['id'])
            service.getTasks(listing['id'],identity)
            updated=service.updatestage(listing['id'],identity)
            print('updated',updated)
            # boolean1,res=service.wait(listing['id'],5)
            # print(boolean1,res)
            # if (boolean1):
            #     updated=service.updatestage(listing['id'],identity)
            #     print('updated',updated)
            # else:
            #     print('couldnt be updated')
            # print('listing[id]',listing['id'])
            # listinges=service.get_listing(listing['id'],identity)
            # print('listinges',listinges)
            # updated=service.updatestage('d281edf0-1bbb-4ab4-8e54-d0db4f0fe8f8',identity)
            # print('updated',updated)

            # service.create_listing(args, identity)

            # print(
            #     "listing created in agent app with id - '{0}' with address- '{1}'".format(
            #         listing["id"], listing["address"]["full_address"]
            #     )
            # )
            # updating bed,bath,car in agent_portal
            # property_details = listing["property_details"]
            # property_details["bedrooms"] = int(lst[3])
            # property_details["bathrooms"] = int(lst[4])
            # property_details["carparks"] = int(lst[5])
            # service.update_property_details(
            #     listing["id"],
            #     {
            #         "appraisals": property_details.get("appraisals", []),
            #         "bathrooms": property_details.get("bathrooms", 0),
            #         "bedrooms": property_details.get("bedrooms", 0),
            #         "carparks": property_details.get("carparks", 0),
            #         "property_type": property_details.get("propertyType", None),
            #         "land_size": property_details.get("landSize", 0),
            #         "internal_sqm": property_details.get("internalSqm", 0),
            #         "external_sqm": property_details.get("externalSqm", 0),
            #         "zoning": property_details.get("zoning", None),
            #     },
            #     identity,
            # )

            # # moving stage from 'opportunity to 'precampaign'
            # item = service.update_stage(
            #     listing["id"], listing["stage_code"].upper(), identity
            # )
            # print('item',item)
            print(1 / 0)

        #     print(
        #         "listing id -'{0}' moved from '{1}' to '{2}'".format(
        #             listing["id"], listing["stage_code"], updated_staged_code
        #         )
        #     )

        #     # creating listing in rex
        #     rex_listing_id = service.create_listing_in_rex(listing["id"])
        #     print("listing created in rex with id - ", rex_listing_id)

        #     # moving agents,bed,bath,car to rex_listing_id
        #     service.update_property_details_in_rex(
        #         rex_listing_id,
        #         identity,
        #         listing["agent_usernames"],
        #         pd["bedrooms"],
        #         pd["bathrooms"],
        #         pd["carparks"],
        #     )

        #     # updating headline and body only if one of the value is given
        #     if lst[6] != None or lst[7] != None:
        #         service.set_headline_and_body_to_rex(rex_listing_id, lst[6], lst[7])

        # print("--------------------------------------------------------\n\n")
