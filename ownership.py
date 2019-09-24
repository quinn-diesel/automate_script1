from openpyxl import load_workbook
import json
import os
import service
import time

EXCEL_PATH = os.getenv("EXCEL_PATH")
SIZE= os.getenv("SIZE")


# PSEUDO

# 1. get the listing id
# 2. swap the agent from ajeya to the one that is in teh list
# 3. 


def main(event, context):

    wb = load_workbook(filename=EXCEL_PATH)
    ws = wb["Sheet1"]

    #want to get every row use 'ws.max_row' instead of 'Limit'
    min_row=2
    max_row=int(SIZE)+int(min_row)-1
    for row in ws.iter_rows(min_row=min_row, min_col=1, max_row=max_row, max_col=15):#want to get every row use 'ws.max_row'
        lst = []
        for cell in row:

            lst.append(cell.value)

        # lst[0]=address value
        # lst[1]=new owner email value

        if lst[0] != None and lst[1] != None:
            print(lst[0],lst[1])
            
            full_address = service.get_property_id_by_address(lst[0])

            get_address = service.get_listing(lst[0])
            print("fulladdress: ", full_address)
            print("get_address: ", get_address)



        

        # print("\n----- fin -----\n")