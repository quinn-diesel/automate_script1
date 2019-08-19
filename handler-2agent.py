from openpyxl import load_workbook
import json
import os
import service
import time

EXCEL_PATH = os.getenv("EXCEL_PATH")
SIZE= os.getenv("SIZE")

def prepare_payload(address, contacts, primary_agent, secondary_agent):
    payload = {
        "arguments": {
            "input": {
                "address": address,
                "contacts": [contacts],
                "primary_agent": primary_agent,
                "secondary_agent": secondary_agent
            }
        },
        "identity": {
            "username": "Google_106600641818143500761",
            "claims": {
                "email": "ajeya@ljx.com.au",
                "custom:office": "corp",
                "custom:manager": "",
                "custom:team": "engineering",
                "cognito:groups": ["agents-group"],
            },
        },
    }
    return payload["arguments"]["input"], payload["identity"]


def main(event, context):

    wb = load_workbook(filename=EXCEL_PATH)
    ws = wb["single_test"]

    #want to get every row use 'ws.max_row' instead of 'Limit'
    min_row=2
    max_row=int(SIZE)+int(min_row)-1
    for row in ws.iter_rows(min_row=min_row, min_col=1, max_row=max_row, max_col=15):#want to get every row use 'ws.max_row'
        lst = []
        for cell in row:

            lst.append(cell.value)

        # lst[0]=address value
        # lst[1]=contact value
        # lst[2]=agent value
        # lst[3]=bed value
        # lst[4]=bath value
        # lst[5]=car value
        # lst[6]=headline value
        # lst[7]=body value
        # list[8]=listed date
        # lst[9]=authority
        # list[10]=price_advertised_as
        # lst[11]=price match as

        if lst[0] != None and lst[1] != None and lst[2] != None:
            print(lst[0],lst[1],lst[2], lst[2])
            args, identity = prepare_payload(
                json.loads(lst[0]),
                json.loads(lst[1]),
                json.loads(lst[2])["primaryAgent"],
                json.loads(lst[2])["secondaryAgent"],
            )

            # creating a listing in agent_portal
            listing = service.create_listing(args, identity)
            print(
                "listing created in agent app with id - '{0}' with address- '{1}'".format(
                    listing["id"], listing["address"]["full_address"]
                )
            )
            # updating bed,bath,car in agent_portal
            property_details = listing["property_details"]
            property_details["bedrooms"] = int(lst[3])
            property_details["bathrooms"] = int(lst[4])
            property_details["carparks"] = int(lst[5])
            pd = service.update_property_details(
                listing["id"],
                {
                    "appraisals": property_details.get("appraisals", []),
                    "bathrooms": property_details.get("bathrooms", 0),
                    "bedrooms": property_details.get("bedrooms", 0),
                    "carparks": property_details.get("carparks", 0),
                    "property_type": property_details.get("propertyType", None),
                    "land_size": property_details.get("landSize", 0),
                    "internal_sqm": property_details.get("internalSqm", 0),
                    "external_sqm": property_details.get("externalSqm", 0),
                    "zoning": property_details.get("zoning", None),
                },
                identity,
            )

            time.sleep(2)

            # # moving stage from 'opportunity to 'precampaign'
            # updated_staged_code = service.update_stage(
            #     listing["id"], identity
            # )
            # print(
            #     "listing id -'{0}' moved from '{1}' to '{2}'".format(
            #         listing["id"], listing["stage_code"], updated_staged_code
            #     )
            # )
            print(
                "listing id -'{0}' moved from '{1}' to '{2}'".format(
                    listing["id"]
                )
            )

            # creating listing in rex
            # rex_listing_id = service.create_listing_in_rex(listing["id"])
            print("listing created in rex with id - ", rex_listing_id)

            # moving agents,bed,bath,car to rex_listing_id
            service.update_property_details_in_rex(
                rex_listing_id,
                identity,
                listing["agent_usernames"],
                pd["bedrooms"],
                pd["bathrooms"],
                pd["carparks"],
            )

            # updating headline and body only if one of the value is given
            # if lst[6] != None or lst[7] != None:
            #     service.set_headline_and_body_to_rex(rex_listing_id, lst[6], lst[7])

            # if lst[8] != None or lst[9] != None:
            #     service.set_listing_details_to_rex(rex_listing_id, date_listed=lst[8], authority=lst[9])

            # if lst[10] != None or lst[11] != None:
            #     service.set_price_to_rex(rex_listing_id, lst[10], lst[11])


        # print("\n--------------------------------------------------------\n")
